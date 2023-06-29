# openpyxl para arquivos Excel xlsx, xlsm, xltx e xltm (instalação)
# Com essa biblioteca será possível ler e escrever dados em células
# específicas, formatar células, inserir gráficos,
# criar fórmulas, adicionar imagens e outros elementos gráficos às suas
# planilhas. Ela é útil para automatizar tarefas envolvendo planilhas do
# Excel, como a criação de relatórios e análise de dados e/ou facilitando a
# manipulação de grandes quantidades de informações.
# Instalação necessária: pip install openpyxl
# Documentação: https://openpyxl.readthedocs.io/en/stable/

from pathlib import Path
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet #Configuração de tipagem 


ROOT_FOLDER = Path(__file__).parent
WORKBOOK_PATH = ROOT_FOLDER / 'workbook.xlsx'

workbook = Workbook() #arquivo
#worksheet:  Worksheet =  woorkbook.active #planilia

# Nome para a planilia
sheet_name = "Minha Planilia"
# Criando uma planília
workbook.create_sheet(sheet_name, 0)
# Selecionou a planília
worksheet:  Worksheet =  workbook.active #planilia
print(workbook.sheetnames)

#Criando cabeçalhos 

worksheet.cell(1,1, "Nome")
worksheet.cell(1,2,"Idade")
worksheet.cell(1,3,"Nota")


students = [
    #nome   idade  nota
    ['João',  14,   5.5],
    ['Maria', 13,   9.7],
    ['Luiz',  15,   8.8],
    ['Alberto', 16,  10],    
]

for i, student_row in enumerate(students, start=2): #Percorrendo as linhas
    for j, student_collumn in enumerate(student_row, start=1): #Percorrendo as colunas
        worksheet.cell(i,j,student_collumn)

#Outro método

for student in students:
    worksheet.append(student) #Coloca a linha de dados da lista na linha de baixo do arquivo 
    #Mesmo resultado !!!

workbook.save(WORKBOOK_PATH)
